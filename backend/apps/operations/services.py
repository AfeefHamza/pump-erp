import hashlib
import csv
import io
import uuid
from decimal import Decimal, InvalidOperation
from django.db import models, transaction
from django.core.exceptions import ValidationError
from django.utils import timezone
from django.core.files.base import ContentFile
import openpyxl

from apps.organizations.models import Organisation, Outlet
from apps.forecourt.models import Tank, Nozzle
from .models import (
    DipCalibrationChart, DipCalibrationPoint, TankCalibrationAssignment,
    OpeningBalanceBatch, NozzleOpeningBalance, TankOpeningBalance,
    NozzleCommissioning, NozzleCommissioningAuditLog
)

def calculate_checksum(file_obj) -> str:
    """
    Calculates SHA256 checksum of a file-like object.
    """
    hasher = hashlib.sha256()
    file_obj.seek(0)
    for chunk in file_obj.chunks():
        hasher.update(chunk)
    file_obj.seek(0)
    return hasher.hexdigest()


def detect_excel_or_csv_columns(file_obj, filename: str) -> dict:
    """
    Reads first few rows of a spreadsheet or CSV and identifies sheet structure and column candidate pairs.
    """
    file_obj.seek(0)
    rows = []
    
    if filename.endswith('.csv'):
        content = file_obj.read().decode('utf-8-sig', errors='ignore')
        reader = csv.reader(io.StringIO(content))
        for i, row in enumerate(reader):
            if i >= 100:
                break
            rows.append(row)
    else:
        wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
        sheet = wb.active
        for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
            if r_idx >= 100:
                break
            rows.append([str(cell) if cell is not None else '' for cell in row])
            
    # Find candidate column pairs (Dip & Volume)
    # Simply look for numeric columns or headers containing 'dip', 'height', 'volume', 'litres', 'mm', 'cm'
    candidate_pairs = []
    headers = []
    if rows:
        headers = rows[0]
        # Basic heuristic to check header text
        for i in range(len(headers)):
            for j in range(len(headers)):
                if i != j:
                    h1 = str(headers[i]).lower()
                    h2 = str(headers[j]).lower()
                    is_h1_dip = any(x in h1 for x in ['dip', 'height', 'mm', 'cm', 'inch'])
                    is_h2_vol = any(x in h2 for x in ['vol', 'litre', 'qty', 'cap'])
                    if is_h1_dip and is_h2_vol:
                        candidate_pairs.append({'dip_idx': i, 'vol_idx': j, 'dip_name': headers[i], 'vol_name': headers[j]})
                        
    return {
        'filename': filename,
        'headers': headers,
        'rows': rows[:20], # limit preview to 20 rows
        'candidate_pairs': candidate_pairs
    }


@transaction.atomic
def import_calibration_chart(organisation, name: str, nominal_capacity: Decimal, lookup_mode: str,
                             original_height_unit: str, file_obj, dip_col_idx: int, vol_col_idx: int,
                             user, **kwargs) -> DipCalibrationChart:
    """
    Parses spreadsheet or CSV, normalizes heights to mm, validates, and imports chart in Draft status.
    """
    filename = file_obj.name
    checksum = calculate_checksum(file_obj)

    # Prevent importing exact duplicate files if needed (optional)
    chart = DipCalibrationChart.objects.create(
        organisation=organisation,
        name=name,
        nominal_capacity=nominal_capacity,
        tank_diameter=kwargs.get('tank_diameter'),
        tank_length=kwargs.get('tank_length'),
        manufacturer_or_source=kwargs.get('manufacturer'),
        source_filename=filename,
        source_file=file_obj,
        source_checksum=checksum,
        original_height_unit=original_height_unit,
        lookup_mode=lookup_mode,
        status=DipCalibrationChart.STATUS_DRAFT,
        imported_by=user
    )

    # Read rows from file
    file_obj.seek(0)
    raw_rows = []
    if filename.endswith('.csv'):
        content = file_obj.read().decode('utf-8-sig', errors='ignore')
        reader = csv.reader(io.StringIO(content))
        raw_rows = list(reader)
    else:
        wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
        sheet = wb.active
        for row in sheet.iter_rows(values_only=True):
            raw_rows.append(row)

    # Parse and normalize values
    points_to_create = []
    seq = 0
    
    # We skip rows that don't have numeric values in the specified columns
    for idx, row in enumerate(raw_rows):
        if not row or len(row) <= max(dip_col_idx, vol_col_idx):
            continue
            
        raw_dip = row[dip_col_idx]
        raw_vol = row[vol_col_idx]
        
        # Try to parse
        try:
            # Strip string if necessary
            if isinstance(raw_dip, str):
                raw_dip = raw_dip.strip()
            if isinstance(raw_vol, str):
                raw_vol = raw_vol.strip()
                
            if raw_dip is None or raw_vol is None or raw_dip == '' or raw_vol == '':
                continue
                
            dip_val = Decimal(str(raw_dip))
            vol_val = Decimal(str(raw_vol))
        except (ValueError, TypeError, ArithmeticError):
            # Skip non-numeric header/footer rows
            continue

        # Normalize height to mm
        if original_height_unit == DipCalibrationChart.UNIT_CM:
            height_mm = dip_val * Decimal('10')
        elif original_height_unit == DipCalibrationChart.UNIT_INCH:
            height_mm = dip_val * Decimal('25.4')
        else:
            height_mm = dip_val

        # Reject negative values
        if height_mm < 0 or vol_val < 0:
            continue

        points_to_create.append(
            DipCalibrationPoint(
                chart=chart,
                height_mm=height_mm,
                volume_litres=vol_val,
                sequence=seq
            )
        )
        seq += 1

    # Bulk create points
    # Sort points by height to ensure sequence corresponds to increasing height
    points_to_create.sort(key=lambda p: p.height_mm)
    for i, p in enumerate(points_to_create):
        p.sequence = i
        
    # Check for duplicate heights
    seen_heights = set()
    unique_points = []
    for p in points_to_create:
        if p.height_mm in seen_heights:
            continue
        seen_heights.add(p.height_mm)
        unique_points.append(p)

    DipCalibrationPoint.objects.bulk_create(unique_points)
    return chart


def validate_chart_monotonicity(chart: DipCalibrationChart) -> tuple[bool, list[str]]:
    """
    Validates that:
    1. Heights strictly increase
    2. Volumes monotonically increase
    3. Volumes match nominal capacity within validation tolerance (e.g. 10%)
    """
    points = list(chart.points.all().order_repr() if hasattr(chart.points.all(), 'order_repr') else chart.points.all().order_by('sequence'))
    if not points:
        return False, ["Chart has no data points."]

    errors = []
    prev_height = Decimal('-1')
    prev_volume = Decimal('-1')

    for p in points:
        if p.height_mm <= prev_height:
            errors.append(f"Height must strictly increase. Found non-increasing height {p.height_mm}mm at sequence {p.sequence}.")
        if p.volume_litres < prev_volume:
            errors.append(f"Volume must increase monotonically. Found volume {p.volume_litres}L at height {p.height_mm}mm which is less than previous volume {prev_volume}L.")
        prev_height = p.height_mm
        prev_volume = p.volume_litres

    # Check nominal capacity matches max volume within 10%
    if points:
        max_vol = points[-1].volume_litres
        diff_pct = abs(max_vol - chart.nominal_capacity) / chart.nominal_capacity
        if diff_pct > Decimal('0.10'):
            errors.append(f"Maximum volume in chart ({max_vol}L) differs from nominal capacity ({chart.nominal_capacity}L) by more than 10% ({diff_pct:.2%}).")

    return len(errors) == 0, errors


@transaction.atomic
def activate_calibration_chart(chart: DipCalibrationChart) -> DipCalibrationChart:
    """
    Validates the calibration chart and changes its status to Active.
    """
    if chart.status == DipCalibrationChart.STATUS_ACTIVE:
        return chart
        
    is_valid, errors = validate_chart_monotonicity(chart)
    if not is_valid:
        raise ValidationError({'detail': f"Chart validation failed: {'; '.join(errors)}"})

    chart.status = DipCalibrationChart.STATUS_ACTIVE
    chart.save()
    return chart


@transaction.atomic
def assign_calibration_chart_to_tank(organisation, outlet, tank: Tank, chart: DipCalibrationChart,
                                     effective_from, user) -> TankCalibrationAssignment:
    """
    Assigns a calibration chart to a tank. Safely closes any overlapping active assignment.
    """
    if tank.organisation_id != organisation.id or chart.organisation_id != organisation.id:
        raise ValidationError("Tank and chart must belong to the same organisation.")
        
    if chart.status != DipCalibrationChart.STATUS_ACTIVE:
        raise ValidationError("Only active calibration charts can be assigned to a tank.")

    # Find the current active assignment and close it
    active_assignments = TankCalibrationAssignment.objects.filter(
        tank=tank,
        effective_to__isnull=True
    )
    for assoc in active_assignments:
        if assoc.effective_from >= effective_from:
            raise ValidationError("A replacement assignment must have an effective date after the current active assignment start date.")
        assoc.effective_to = effective_from
        assoc.save()

    # Create new assignment
    return TankCalibrationAssignment.objects.create(
        organisation=organisation,
        outlet=outlet,
        tank=tank,
        chart=chart,
        effective_from=effective_from,
        assigned_by=user
    )


def convert_dip_to_volume(tank: Tank, measured_height: Decimal, input_unit: str, measured_at=None) -> dict:
    """
    Finds the calibration chart assignment active at measured_at, normalizes the measured height,
    looks up/interpolates volume, and returns volume conversion details.
    """
    if not measured_at:
        measured_at = timezone.now()

    # Find active assignment
    assignment = TankCalibrationAssignment.objects.filter(
        tank=tank,
        effective_from__lte=measured_at
    ).filter(
        models.Q(effective_to__isnull=True) | models.Q(effective_to__gt=measured_at)
    ).select_related('chart').first()

    if not assignment:
        raise ValidationError(f"No active calibration assignment found for tank {tank.name} at {measured_at.strftime('%Y-%m-%d %H:%M')}.")

    chart = assignment.chart
    
    # Normalize input height to mm
    if input_unit == DipCalibrationChart.UNIT_CM:
        height_mm = measured_height * Decimal('10')
    elif input_unit == DipCalibrationChart.UNIT_INCH:
        height_mm = measured_height * Decimal('25.4')
    else:
        height_mm = measured_height

    if height_mm < 0:
        raise ValidationError("Height measurement cannot be negative.")

    # Fetch all points sorted by height
    points = list(chart.points.all().order_by('height_mm'))
    if not points:
        raise ValidationError("Calibration chart contains no lookup points.")

    # Exact lookup check
    exact_match = next((p for p in points if p.height_mm == height_mm), None)
    if exact_match:
        return {
            'volume': exact_match.volume_litres,
            'chart': chart,
            'assignment': assignment,
            'method': TankOpeningBalance.METHOD_EXACT,
            'surrounding_points': None
        }

    # Interpolation checks
    if chart.lookup_mode == DipCalibrationChart.LOOKUP_EXACT:
        raise ValidationError(f"Exact match lookup required, but no point found for height {height_mm}mm.")

    # Linear interpolation: find lower and upper bounds
    lower_point = None
    upper_point = None
    
    for p in points:
        if p.height_mm < height_mm:
            lower_point = p
        elif p.height_mm > height_mm:
            upper_point = p
            break

    if not lower_point or not upper_point:
        # Check boundary edge cases
        # Allow exact zero height to return zero volume if no lower point exists and height is 0
        if height_mm == 0:
            return {
                'volume': Decimal('0'),
                'chart': chart,
                'assignment': assignment,
                'method': TankOpeningBalance.METHOD_EXACT,
                'surrounding_points': None
            }
        raise ValidationError(f"Extrapolation prohibited. Measured height {height_mm}mm is outside calibration chart range [{points[0].height_mm}mm - {points[-1].height_mm}mm].")

    # Perform linear interpolation
    h_diff = upper_point.height_mm - lower_point.height_mm
    v_diff = upper_point.volume_litres - lower_point.volume_litres
    pct = (height_mm - lower_point.height_mm) / h_diff
    volume = lower_point.volume_litres + (v_diff * pct)

    return {
        'volume': volume.quantize(Decimal('0.0001')),
        'chart': chart,
        'assignment': assignment,
        'method': TankOpeningBalance.METHOD_INTERPOLATE,
        'surrounding_points': {
            'lower': {'height': lower_point.height_mm, 'volume': lower_point.volume_litres},
            'upper': {'height': upper_point.height_mm, 'volume': upper_point.volume_litres}
        }
    }


@transaction.atomic
def create_opening_balance_batch(organisation, outlet, effective_at, user, notes=None) -> OpeningBalanceBatch:
    """
    Creates a new opening balance batch for preparing. Ensures only one confirmed batch can exist.
    """
    if OpeningBalanceBatch.objects.filter(outlet=outlet, status=OpeningBalanceBatch.STATUS_CONFIRMED).exists():
        raise ValidationError("This outlet already has a confirmed opening balance batch. Cannot create another.")

    # Delete any existing preparing batches to start fresh if needed, or raise warning
    OpeningBalanceBatch.objects.filter(outlet=outlet, status=OpeningBalanceBatch.STATUS_PREPARING).delete()

    return OpeningBalanceBatch.objects.create(
        organisation=organisation,
        outlet=outlet,
        effective_at=effective_at,
        status=OpeningBalanceBatch.STATUS_PREPARING,
        created_by=user,
        notes=notes
    )


@transaction.atomic
def set_nozzle_opening_balance(batch: OpeningBalanceBatch, nozzle: Nozzle, totalizer_reading: Decimal, notes=None) -> NozzleOpeningBalance:
    """
    Sets the opening totalizer reading for a nozzle in the batch.
    """
    if batch.status == OpeningBalanceBatch.STATUS_CONFIRMED:
        raise ValidationError("Cannot modify confirmed opening balances.")
        
    return NozzleOpeningBalance.objects.update_or_create(
        batch=batch,
        nozzle=nozzle,
        defaults={
            'totalizer_reading': totalizer_reading,
            'notes': notes
        }
    )


@transaction.atomic
def set_tank_opening_balance(batch: OpeningBalanceBatch, tank: Tank, book_quantity: Decimal,
                             physical_quantity: Decimal, **kwargs) -> TankOpeningBalance:
    """
    Sets the opening stock levels for a tank in the batch.
    """
    if batch.status == OpeningBalanceBatch.STATUS_CONFIRMED:
        raise ValidationError("Cannot modify confirmed opening balances.")

    return TankOpeningBalance.objects.update_or_create(
        batch=batch,
        tank=tank,
        defaults={
            'book_quantity': book_quantity,
            'physical_quantity': physical_quantity,
            'raw_dip_value': kwargs.get('raw_dip_value'),
            'raw_dip_unit': kwargs.get('raw_dip_unit'),
            'calibration_assignment': kwargs.get('calibration_assignment'),
            'density': kwargs.get('density'),
            'conversion_method': kwargs.get('conversion_method', TankOpeningBalance.METHOD_MANUAL),
            'manual_quantity_reason': kwargs.get('manual_quantity_reason'),
            'notes': kwargs.get('notes')
        }
    )


def preview_opening_balance_confirmation(batch: OpeningBalanceBatch) -> dict:
    """
    Gathers validation statistics and items to show a readiness preview for opening balance confirmation.
    """
    outlet = batch.outlet
    
    # Active dispensers/nozzles/tanks
    active_nozzles = list(Nozzle.objects.filter(outlet=outlet, status=Nozzle.STATUS_ACTIVE))
    active_tanks = list(Tank.objects.filter(outlet=outlet, status=Tank.STATUS_ACTIVE))
    
    nozzle_balances = {nb.nozzle_id: nb for nb in batch.nozzle_balances.all()}
    tank_balances = {tb.tank_id: tb for tb in batch.tank_balances.all()}
    
    missing_nozzles = []
    nozzle_details = []
    for nozzle in active_nozzles:
        bal = nozzle_balances.get(nozzle.id)
        if not bal:
            missing_nozzles.append(nozzle)
        nozzle_details.append({
            'nozzle_id': nozzle.id,
            'nozzle_code': nozzle.code,
            'reading': bal.totalizer_reading if bal else None,
            'is_configured': bal is not None
        })

    missing_tanks = []
    tank_details = []
    for tank in active_tanks:
        bal = tank_balances.get(tank.id)
        if not bal:
            missing_tanks.append(tank)
        tank_details.append({
            'tank_id': tank.id,
            'tank_code': tank.code,
            'book_quantity': bal.book_quantity if bal else None,
            'physical_quantity': bal.physical_quantity if bal else None,
            'is_configured': bal is not None
        })

    ready = (len(missing_nozzles) == 0 and len(missing_tanks) == 0)
    
    return {
        'ready': ready,
        'batch_id': batch.id,
        'effective_at': batch.effective_at,
        'missing_nozzle_count': len(missing_nozzles),
        'missing_tank_count': len(missing_tanks),
        'nozzles': nozzle_details,
        'tanks': tank_details
    }


@transaction.atomic
def confirm_opening_balance_batch(batch: OpeningBalanceBatch, user) -> OpeningBalanceBatch:
    """
    Atomically confirms the complete opening balance batch. Enforces completeness rules.
    """
    if batch.status == OpeningBalanceBatch.STATUS_CONFIRMED:
        return batch

    preview = preview_opening_balance_confirmation(batch)
    if not preview['ready']:
        raise ValidationError(
            f"Cannot confirm batch. Missing readings for {preview['missing_nozzle_count']} nozzles and {preview['missing_tank_count']} tanks."
        )

    # Perform confirmation
    batch.status = OpeningBalanceBatch.STATUS_CONFIRMED
    batch.confirmed_by = user
    batch.confirmed_at = timezone.now()
    batch.save()
    return batch


@transaction.atomic
def commission_nozzle(
    organisation,
    outlet,
    nozzle,
    initial_totalizer,
    effective_at,
    reason: str,
    notes: str | None = None,
    actor=None,
    activate: bool = True
) -> NozzleCommissioning:
    """
    Establishes the authoritative initial totalizer for a newly added nozzle.
    1. Lock the nozzle and relevant starting-reading records.
    2. Validate tenant and outlet consistency.
    3. Confirm the nozzle has no existing authoritative starting record that makes commissioning invalid.
    4. Confirm it has no closed-shift history.
    5. Validate the effective timestamp.
    6. Create the immutable commissioning record.
    7. Recalculate outlet readiness.
    8. Add an append-only audit event.
    9. Complete atomically.
    """
    from apps.shifts.models import OperationalShift, ShiftNozzleMeter
    from .selectors import check_outlet_operational_readiness

    if isinstance(nozzle, (str, uuid.UUID)):
        try:
            nozzle = Nozzle.objects.select_for_update().select_related(
                'dispenser', 'tank', 'tank__product', 'outlet', 'organisation'
            ).get(id=nozzle)
        except Nozzle.DoesNotExist:
            raise ValidationError("Nozzle does not exist.")
    else:
        # Row-lock nozzle
        nozzle = Nozzle.objects.select_for_update().select_related(
            'dispenser', 'tank', 'tank__product', 'outlet', 'organisation'
        ).get(id=nozzle.id)

    # Validate tenant and outlet consistency
    if outlet.organisation_id != organisation.id:
        raise ValidationError("Outlet must belong to the specified organisation.")

    if nozzle.outlet_id != outlet.id:
        raise ValidationError("Nozzle must belong to the specified outlet.")

    if nozzle.organisation_id != organisation.id:
        raise ValidationError("Nozzle must belong to the specified organisation.")

    # Status check
    if nozzle.status != Nozzle.STATUS_ACTIVE:
        if activate:
            nozzle.status = Nozzle.STATUS_ACTIVE
            nozzle.save(update_fields=['status', 'updated_at'])
        else:
            raise ValidationError("Nozzle is not active. It must be active or explicitly activated through this workflow.")

    # Confirm the nozzle has no existing authoritative starting record
    has_opening_balance = NozzleOpeningBalance.objects.filter(
        batch__outlet=outlet,
        batch__status=OpeningBalanceBatch.STATUS_CONFIRMED,
        nozzle=nozzle
    ).exists()
    if has_opening_balance:
        raise ValidationError(f"Nozzle '{nozzle.code}' already has a confirmed opening balance. Commissioning is not permitted.")

    if NozzleCommissioning.objects.filter(nozzle=nozzle).exists():
        raise ValidationError(f"Nozzle '{nozzle.code}' has already been commissioned. Ordinary commissioning is allowed only once per nozzle.")

    # Confirm it has no closed-shift history
    has_closed_shift = ShiftNozzleMeter.objects.filter(
        nozzle=nozzle,
        shift__status=OperationalShift.STATUS_CLOSED
    ).exists()
    if has_closed_shift:
        raise ValidationError(f"Nozzle '{nozzle.code}' has historical closed-shift meter data. Ordinary commissioning is not permitted.")

    # Validate totalizer
    if initial_totalizer is None or initial_totalizer == '':
        raise ValidationError("Initial totalizer reading is required.")

    try:
        totalizer_dec = Decimal(str(initial_totalizer))
    except (InvalidOperation, TypeError, ValueError):
        raise ValidationError("Initial totalizer must be a valid numeric decimal.")

    if totalizer_dec < Decimal('0.000'):
        raise ValidationError("Initial totalizer cannot be negative.")

    if not effective_at:
        raise ValidationError("Effective date and time is required.")

    if timezone.is_naive(effective_at):
        effective_at = timezone.make_aware(effective_at)

    if not reason or not str(reason).strip():
        raise ValidationError("A mandatory reason is required for nozzle commissioning.")

    clean_reason = str(reason).strip()
    clean_notes = str(notes).strip() if notes and str(notes).strip() else None

    # Create immutable commissioning record
    commissioning = NozzleCommissioning(
        organisation=organisation,
        outlet=outlet,
        nozzle=nozzle,
        effective_at=effective_at,
        initial_totalizer=totalizer_dec,
        reason=clean_reason,
        notes=clean_notes,
        commissioned_by=actor,
        dispenser_code_snapshot=nozzle.dispenser.code,
        nozzle_code_snapshot=nozzle.code,
        product_id_snapshot=nozzle.tank.product.id,
        product_name_snapshot=nozzle.tank.product.name,
    )
    commissioning.full_clean()
    commissioning.save()

    # Add append-only audit event
    audit_log = NozzleCommissioningAuditLog(
        organisation=organisation,
        outlet=outlet,
        commissioning=commissioning,
        nozzle=nozzle,
        event_type='nozzle_commissioned',
        actor=actor,
        reason=clean_reason,
        metadata={
            'nozzle_id': str(nozzle.id),
            'nozzle_code': nozzle.code,
            'dispenser_id': str(nozzle.dispenser.id),
            'dispenser_code': nozzle.dispenser.code,
            'product_id': str(nozzle.tank.product.id),
            'product_name': nozzle.tank.product.name,
            'initial_totalizer': str(totalizer_dec),
            'effective_at': effective_at.isoformat(),
            'notes': clean_notes,
        }
    )
    audit_log.full_clean()
    audit_log.save()

    # Recalculate outlet readiness
    check_outlet_operational_readiness(outlet)

    return commissioning


@transaction.atomic
def bulk_commission_nozzles(
    organisation,
    outlet,
    items: list[dict],
    effective_at,
    common_reason: str,
    actor,
    activate: bool = True
) -> list[NozzleCommissioning]:
    """
    Atomically commissions multiple nozzles together.
    Validates all rows; if one row fails, none are saved.
    """
    if not items:
        raise ValidationError("At least one nozzle must be provided for bulk commissioning.")

    if not common_reason or not str(common_reason).strip():
        raise ValidationError("A common mandatory reason is required for bulk commissioning.")

    if not effective_at:
        raise ValidationError("A common effective date and time is required for bulk commissioning.")

    if timezone.is_naive(effective_at):
        effective_at = timezone.make_aware(effective_at)

    clean_reason = str(common_reason).strip()

    seen_nozzle_ids = set()
    for idx, item in enumerate(items):
        raw_nozzle = item.get('nozzle') or item.get('nozzle_id')
        nozzle_id = str(raw_nozzle.id if hasattr(raw_nozzle, 'id') else raw_nozzle)
        if nozzle_id in seen_nozzle_ids:
            raise ValidationError(f"Row {idx + 1}: Duplicate nozzle in bulk commissioning list.")
        seen_nozzle_ids.add(nozzle_id)

    created_records = []
    for idx, item in enumerate(items):
        nozzle = item.get('nozzle') or item.get('nozzle_id')
        reading = item.get('initial_totalizer')
        row_notes = item.get('notes')

        if reading is None or reading == '':
            raise ValidationError(f"Row {idx + 1}: Initial totalizer is required.")

        try:
            reading_dec = Decimal(str(reading))
        except (InvalidOperation, TypeError, ValueError):
            raise ValidationError(f"Row {idx + 1}: Initial totalizer must be a valid numeric decimal.")

        if reading_dec < Decimal('0.000'):
            raise ValidationError(f"Row {idx + 1}: Initial totalizer cannot be negative.")

        comm = commission_nozzle(
            organisation=organisation,
            outlet=outlet,
            nozzle=nozzle,
            initial_totalizer=reading_dec,
            effective_at=effective_at,
            reason=clean_reason,
            notes=row_notes,
            actor=actor,
            activate=activate
        )
        created_records.append(comm)

    return created_records

